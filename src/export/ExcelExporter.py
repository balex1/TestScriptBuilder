# -*- coding: utf-8 -*-
"""
Created on Fri Nov 27 11:06:25 2015

@author: Alex
"""

import xml.etree.ElementTree as ET
from openpyxl import Workbook
import platform
if platform.system() == 'Windows':
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    import openpyxl.utils as Utils
import sqlite3 as lite
import os
from ObjectClasses import _Workflow, _KeyAction, _InputParameter

class TemplateReader():
    #Read the XML Template and generate SQL Queries based on it
    #Pass to the writer each segment individually

    def __init__(self, db_path):
        self.wb = Workbook()
        orig_sheet = self.wb.get_active_sheet()
        if platform.system() == 'Windows':
            self.wb.create_sheet('Header', 0)
        else:
            self.wb.create_sheet(0, 'Header')
        self.wb.remove_sheet(orig_sheet)
        self.db_path=db_path
        
        if platform.system() == 'Windows':
        
            #Base Styles
            self.base_font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
            self.base_fill = PatternFill(fill_type=None, start_color='FFFFFFFF', end_color='FF000000')
            self.base_border = Border(left=Side(border_style=None, color='FF000000'), right=Side(border_style=None, color='FF000000'),\
                top=Side(border_style=None, color='FF000000'), bottom=Side(border_style=None, color='FF000000'), diagonal=Side(border_style=None, color='FF000000'),\
                    diagonal_direction=0, outline=Side(border_style=None, color='FF000000'), vertical=Side(border_style=None, color='FF000000'), horizontal=Side(border_style=None, color='FF000000'))
            self.base_alignment=Alignment(horizontal='general', vertical='bottom', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
            self.base_number_format = 'General'
            
            #Base Header Styles
            self.header_font = Font(name='Calibri', size=13, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
            self.header_fill = PatternFill(fill_type=None, start_color='FFFFFFFF', end_color='FF000000')
            self.header_border = Border(left=Side(border_style=None, color='FF000000'), right=Side(border_style=None, color='FF000000'),\
                top=Side(border_style=None, color='FF000000'), bottom=Side(border_style=None, color='FF000000'), diagonal=Side(border_style=None, color='FF000000'),\
                    diagonal_direction=0, outline=Side(border_style='thin', color='FF000000'), vertical=Side(border_style=None, color='FF000000'), horizontal=Side(border_style=None, color='FF000000'))
            self.header_alignment=Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
            self.header_number_format = 'General'

    def select_files_in_folder(self, dir, ext):
        for file in os.listdir(dir):
            if file.endswith('.%s' % ext):
                yield os.path.join(dir, file)
                
    def generate_parameter_list(self, xml_path):
        
        #Param list to hold input parameters
        param_list=[]
        #Read the XML Template
        self.tree = ET.parse(xml_path)
        self.root = self.tree.getroot()
        for child in self.root:
            if child.tag == 'InputParameters':
                for value in child:
                    #Add each value to the parameter list
                    param_list.append(value.text)
        return param_list
        
    def execute_workflow_export(self, flow, worksheet, row, column):
        #Export the _Workflow object to the specified row and column on the excel sheet
        pass
        
    def generate_workflow_export(self, workflow_name, testscript, project, client, worksheet, row, column):
        
        #The number of rows in the workflow to be returned
        num_rows = 0
        
        #Find the workflow
        self.cur.execute('select wf.id, wf.name from workflow wf left join testscript ts on ts.id = wf.testscriptid) left join project p on ts.projectid = p.id) left join client c on p.clientid = c.id where wf.name = %s and ts.name = %s and p.name = %s and c.name = %s order by w.id;' % (workflow_name, testscript, project, client))
        flow = self.cur.fetchone()
        
        workflow = _Workflow()
        workflow.name = flow[1]
        workflow.id = flow[0]
        
        #Find the Key Actions for the workflow
        self.cur.execute('select ka.id, ka.name, ka.description, ka.custom, wfa.expectedresult, wfa.notes from ((workflowaction wfa left join keyaction ka on wfa.keyactionid = ka.id) left join workflow w on wfa.workflowid = w.id) where w.id = %s;' % (workflow[0]))
        keyactions = self.cur.fetchall()
        
        for action in keyactions:
            
            keyaction = _KeyAction()
            keyaction.id = action[0]
            keyaction.name = action[1]
            keyaction.description = action[2]
            if action[3] == 0:
                keyaction.custom = False
            else:
                keyaction.custom = True
            keyaction.expected_result = action[4]
            keyaction.notes = action[5]
            workflow.add_keyaction(keyaction)
            
            #Find the Input Parameters for the Key Action
            self.cur.execute('select ip.id, ip.name, wp.value from ((inputparameter ip left join keyaction ka on ip.keyactionid = ka.id) left join workflowparam wp on wp.inputparamid = ip.id) where ka.id = %s;' % (action[0]))
            inputparameters = self.cur.fetchall()
            for param in inputparameters:
                input_parameter = _InputParameter()
                input_parameter.id = param[0]
                input_parameter.name = param[1]
                input_parameter.value = param[2]
                keyaction.add_inputparameter(input_parameter)
                num_rows+=1
                
        #TODO - Write the _Workflow object to the Excel Sheet
        self.execute_workflow_export(workflow, worksheet, row, column)
            
        return num_rows
    
    #Takes in the XML path of the template file
    #params is a list of the input parameters for the template
    #Look for '?' in the xml values and replace it with the next up param
    def translate_template(self, xml_path, params):
        
        #Counter for the input parameters
        param_counter = 0
        
        #Counter for the segments within pages
        segment_counter = 0
        
        #Wild Card Counter
        wc_counter = 0
        
        #Clear the sheets out of the workbook
        sn = self.wb.get_sheet_names()
        for s in sn:
            self.wb.remove_sheet(self.wb.get_sheet_by_name(s))
        self.wb.create_sheet('Header', 0)
        
        #Connect to the DB
        self.con = lite.connect(self.db_path)
        self.cur = self.con.cursor()
        
        with self.con:
        
            #Read the XML Template
            self.tree = ET.parse(xml_path)
            self.root = self.tree.getroot()
            header_ws = self.wb.get_sheet_by_name('Header')
            for child in self.root:
                if child.tag == 'Header':
                    for row in child:
                        #Process each header row
                        for element in row:
                            #Process each element
                            #We expect two property values:
                            #cell, the first cell of the merge
                            #end_cell, the last cell in the merge
                            if '?' not in element.text:
                                header_ws[element.attrib['start_cell']] = element.text
                            #If a wildcard is encountered, we need to replace it with the
                            #correct parameter
                            else:
                                text = list(element.text)
                                wc_counter = 0
                                for i in text:
                                    if i == '?':
                                        text[wc_counter] = params[int(text[wc_counter + 1])]
                                        print('Parameter %s used' % (params[int(text[wc_counter + 1])]))
                                        del text[wc_counter + 1]
                                        param_counter+=1
                                    wc_counter+=1
                                header_ws[element.attrib['start_cell']] = ''.join(text)
                            if platform.system() == 'Windows':
                                header_ws[element.attrib['start_cell']].font = self.header_font
                                header_ws[element.attrib['start_cell']].fill = self.header_fill
                                header_ws[element.attrib['start_cell']].border = self.header_border
                                header_ws[element.attrib['start_cell']].alignment = self.header_alignment
                                header_ws[element.attrib['start_cell']].number_format = self.header_number_format
                            header_ws.merge_cells('%s:%s' % (element.attrib['start_cell'], element.attrib['end_cell']))
                            print('Header element placed')
                elif child.tag == 'Body':
                    #Process the body segment
                    for page in child:
                        #Process each page
                        segment_counter = 0
                        body_ws = self.wb.create_sheet(page.attrib['name'])
                        for segment in page:
                            if segment.tag == 'TestScriptSteps':
                                #Execute the specialized Test Script Steps Export
                                #This segment needs to be hardcoded due to the ability to construct
                                #nonlinear workflows.
                                #Here, we process a full testscript and output each workflow in an optimized state
                            
                                #We expect the first input parameter to be Test Script, then Project, then Client, 
                                #After those, others can be used and added to the template
                            
                                start_cell = segment.attrib['cell']
                                
                                #Set the parameter counter so that it doesn't break after running this
                                param_counter+=3
                                
                                #Find the workflows associated with the test script
                                self.cur.execute('select wf.id, wf.name from workflow wf left join testscript ts on ts.id = wf.testscriptid) left join project p on ts.projectid = p.id) left join client c on p.clientid = c.id where ts.name = %s and p.name = %s and c.name = %s order by w.id;' % (params[0], params[1], params[2]))
                                workflows = self.cur.fetchall()
                                
                                #Iterate over the cells
                                col = Utils.column_index_from_string(start_cell[0])
                                row = int(float(start_cell[1]))
                                
                                flow_counter = 0
                                
                                for workflow in workflows:
                                    w_row = row + flow_counter
                                    w_col = col
                                    flow_counter += self.generate_workflow_export(workflow[1], params[0], params[1], params[2], body_ws, w_row, w_col)
                            elif segment.tag == 'WorkflowSteps':
                                #Execute a single workflow export
                            
                                start_cell = segment.attrib['cell']
                                
                                col = Utils.column_index_from_string(start_cell[0])
                                row = int(float(start_cell[1]))
                            
                                #We expect the first input parameter to be Test Script, then Project, then Client, then Workflow
                                #After those, others can be used and added to the template
                                self.generate_workflow_export(params[3], params[0], params[1], params[2], row, col)
                            else:
                                for child in segment:
                                    if child.tag == 'Title':
                                        if '?' not in child.text:
                                            body_ws[segment.attrib['cell']] = child.text
                                        #If a wildcard is encountered, we need to replace it with the
                                        #correct parameter
                                        else:
                                            text = list(child.text)
                                            wc_counter = 0
                                            for i in text:
                                                if i == '?':
                                                    text[wc_counter] = params[int(text[wc_counter + 1])]
                                                    print('Parameter %s used' % (params[int(text[wc_counter + 1])]))
                                                    del text[wc_counter + 1]
                                                    param_counter+=1
                                                wc_counter+=1
                                            body_ws[segment.attrib['cell']] = ''.join(text)
                                        if platform.system() == 'Windows':
                                            body_ws[segment.attrib['cell']].font = self.header_font
                                            body_ws[segment.attrib['cell']].fill = self.header_fill
                                            body_ws[segment.attrib['cell']].border = self.header_border
                                            body_ws[segment.attrib['cell']].alignment = self.header_alignment
                                            body_ws[segment.attrib['cell']].number_format = self.header_number_format
                                        print('Data Title element %s placed in cell %s' % (child.text, segment.attrib['cell']))
                                        segment_counter+=1
                                    elif child.tag == 'Header':
                                        i=0
                                        for column in child:
                                            #Place the column header for each query column
                                            cell = Utils.coordinate_from_string(segment.attrib['cell'])
                                            col = Utils.column_index_from_string(cell[0])
                                            body_ws['%s%s' % (Utils.get_column_letter(col+i), 1 + segment_counter)] = column.text
                                            if platform.system() == 'Windows':
                                                body_ws['%s%s' % (Utils.get_column_letter(col+i), 1 + segment_counter)].font = self.base_font
                                                body_ws['%s%s' % (Utils.get_column_letter(col+i), 1 + segment_counter)].fill = self.base_fill
                                                body_ws['%s%s' % (Utils.get_column_letter(col+i), 1 + segment_counter)].border = self.base_border
                                                body_ws['%s%s' % (Utils.get_column_letter(col+i), 1 + segment_counter)].alignment = self.base_alignment
                                                body_ws['%s%s' % (Utils.get_column_letter(col+i), 1 + segment_counter)].number_format = self.base_number_format
                                            print('Data Header element %s placed in cell %s%s' % (column.text, Utils.get_column_letter(col+i), 2))
                                            i+=1
                                        segment_counter+=1
                                    elif child.tag == 'Query':
                                        #Execute the query and place the results into the page
                                        if '?' not in child.text:
                                            self.cur.execute(child.text)
                                        #If a wildcard is encountered, we need to replace it with the
                                        #correct parameter
                                        else:
                                            text = list(child.text)
                                            wc_counter = 0
                                            for i in text:
                                                if i == '?':
                                                    text[wc_counter] = params[int(text[wc_counter + 1])]
                                                    print('Parameter %s used' % (params[int(text[wc_counter + 1])]))
                                                    del text[wc_counter + 1]
                                                    param_counter+=1
                                                wc_counter+=1
                                            self.cur.execute(''.join(text))
                                        data = self.cur.fetchall()
                                        print('query %s executed' % (child.text))
                                        i=3
                                        for row in data:
                                            j=0
                                            segment_counter+=1
                                            #Place the data into the report
                                            for e in row:
                                                cell = Utils.coordinate_from_string(segment.attrib['cell'])
                                                col = Utils.column_index_from_string(cell[0])
                                                body_ws['%s%s' % (Utils.get_column_letter(col+j), i)] = e
                                                if platform.system() == 'Windows':
                                                    body_ws['%s%s' % (Utils.get_column_letter(col+j), i)].font = self.base_font
                                                    body_ws['%s%s' % (Utils.get_column_letter(col+j), i)].fill = self.base_fill
                                                    body_ws['%s%s' % (Utils.get_column_letter(col+j), i)].border = self.base_border
                                                    body_ws['%s%s' % (Utils.get_column_letter(col+j), i)].alignment = self.base_alignment
                                                    body_ws['%s%s' % (Utils.get_column_letter(col+j), i)].number_format = self.base_number_format
                                                print('Data Element %s placed in column %s%s' % (e, Utils.get_column_letter(col+i), j))
                                                j+=1
                                            i+=1
            self.wb.save('Export.xlsx')